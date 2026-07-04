"""Excel exporter.

Generates the final Excel workbook containing normalized orders.
Applies formatting, safe numeric exports (Decimal precision), 
and mitigates CSV/Excel formula injection.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pandas as pd

from deliverect_sync.config import AppSettings
from deliverect_sync.logging_config import get_logger
from deliverect_sync.security.pii import PIIFieldEncryption
from deliverect_sync.storage.database import DatabaseManager

logger = get_logger("excel_exporter")


def sanitize_formula(value: Any) -> Any:
    """Sanitize strings to prevent CSV/Excel formula injection.
    
    Prepends a single quote to strings starting with formula characters.
    Does not touch numeric or boolean types.
    """
    if not isinstance(value, str):
        return value
        
    if value.startswith(("=", "+", "-", "@", "\t", "\r")):
        return f"'{value}"
        
    return value


class ExcelExporter:
    """Exports synchronized orders to a formatted Excel workbook."""

    def __init__(self, settings: AppSettings, db: DatabaseManager) -> None:
        self._settings = settings
        self._db = db
        self._pii = PIIFieldEncryption()

    def export(self, run_id: str) -> Path:
        """Generate the Excel export for a given sync run.

        Args:
            run_id: ID of the sync run (for logging context, but exports all).

        Returns:
            Path to the generated Excel file.
        """
        logger.info("Generating Excel export for run %s", run_id[:8])

        # pyrefly: ignore [missing-attribute]
        self._settings.output_dir.mkdir(parents=True, exist_ok=True)

        orders = self._db.get_all_orders()
        if not orders:
            logger.warning("No orders found to export")
            return self._create_empty_export()

        # Decrypt PII where applicable
        self._decrypt_pii(orders)

        df_orders = pd.DataFrame(orders)

        # Ensure correct Data Types
        self._format_dataframe(df_orders)

        timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"deliverect_sync_{timestamp}.xlsx"
        # pyrefly: ignore [missing-attribute]
        output_path = Path(self._settings.output_dir) / filename

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self._write_sheet(writer, df_orders, "Orders")

            if self._settings.export.split_items_into_separate_rows:
                items = self._db.get_all_order_items()
                if items:
                    df_items = pd.DataFrame(items)
                    self._format_dataframe(df_items)
                    self._write_sheet(writer, df_items, "Items")
                    
            run = self._db.get_last_run()
            if run:
                summary_data = {
                    "Run ID": [run.id],
                    "Started At": [run.started_at.isoformat() if run.started_at else None],
                    "Finished At": [run.finished_at.isoformat() if run.finished_at else None],
                    # pyrefly: ignore [redundant-condition]
                    "Result": [run.result.value if run.result else None],
                    "Imported Rows": [run.imported_rows],
                    "New Orders": [run.new_orders],
                    "Updated Orders": [run.updated_orders],
                    "Rejected Rows": [run.rejected_rows],
                }
                df_summary = pd.DataFrame(summary_data)
                self._write_sheet(writer, df_summary, "Summary")

        logger.info("Excel export completed: %s", output_path)
        return output_path

    def _format_dataframe(self, df: pd.DataFrame) -> None:
        """Format DataFrame columns for Excel output.
        
        1. Parse dates safely.
        2. Convert SQLite text decimals to actual Python Decimals.
        3. Apply formula injection protection to string columns.
        """
        date_cols = ["pickup_time_utc", "first_seen_at", "last_seen_at"]
        for col in date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], format="mixed", errors="coerce").dt.tz_localize(None)

        decimal_cols = ["order_total", "subtotal", "tax_total", "delivery_fee", 
                        "service_charge", "tip", "discount_total", "unit_price", "line_total"]
        
        def safe_decimal(x: Any) -> Decimal | None:
            if pd.isna(x):
                return None
            s = str(x).strip()
            if not s:
                return None
            try:
                return Decimal(s)
            except InvalidOperation:
                return None

        for col in decimal_cols:
            if col in df.columns:
                # pyrefly: ignore [no-matching-overload]
                df[col] = df[col].apply(safe_decimal)
                
        # Apply formula injection protection
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(sanitize_formula)

    def _decrypt_pii(self, orders: list[dict[str, Any]]) -> None:
        """Decrypt PII fields in the order records."""
        for order in orders:
            if self._settings.privacy.include_customer_name and order.get("customer_name_encrypted"):
                order["Customer Name"] = self._pii.decrypt_field(order["customer_name_encrypted"])
            if self._settings.privacy.include_customer_phone:
                if order.get("customer_phone_e164_encrypted"):
                    order["Customer Phone (E164)"] = self._pii.decrypt_field(order["customer_phone_e164_encrypted"])
                if order.get("customer_phone_original_encrypted"):
                    order["Customer Phone (Original)"] = self._pii.decrypt_field(order["customer_phone_original_encrypted"])
            if self._settings.privacy.include_customer_email and order.get("customer_email_encrypted"):
                order["Customer Email"] = self._pii.decrypt_field(order["customer_email_encrypted"])
            if self._settings.privacy.include_delivery_address and order.get("delivery_address_encrypted"):
                order["Delivery Address"] = self._pii.decrypt_field(order["delivery_address_encrypted"])

            # Clean up raw encrypted fields from output
            for field in ["customer_name_encrypted", "customer_phone_original_encrypted", "customer_phone_e164_encrypted", 
                          "customer_email_encrypted", "delivery_address_encrypted"]:
                if field in order:
                    del order[field]

    def _write_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
        """Write a DataFrame to a sheet with basic formatting."""
        internal_cols = ["id", "source_row_hash", "source_file_id"]
        df_clean = df.drop(columns=[c for c in internal_cols if c in df.columns])

        df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # Auto-adjust columns width
        for i, col in enumerate(df_clean.columns):
            column_len = max(
                df_clean[col].astype(str).map(len).max() if not df_clean[col].empty else 0,
                # pyrefly: ignore [unnecessary-type-conversion]
                len(str(col))
            )
            adjusted_width = min(max(column_len + 2, 10), 50)
            
            # Excel columns are 1-indexed for some libraries, but openpyxl uses letters A, B, C etc or 1-indexed ints
            # Openpyxl allows `chr(65+i)` for up to Z, but fails for AA. Better to use get_column_letter
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    def _create_empty_export(self) -> Path:
        """Create an empty Excel file with headers only."""
        timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"deliverect_sync_{timestamp}_empty.xlsx"
        # pyrefly: ignore [missing-attribute]
        output_path = Path(self._settings.output_dir) / filename
        
        df = pd.DataFrame(columns=["Notice"])
        df.loc[0] = ["No orders found for this export run."]
        
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Orders", index=False)
            
        return output_path
